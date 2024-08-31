import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from tkinter import Tk, filedialog, messagebox, Button
import time

# Configure logging
logging.basicConfig(filename='scrapify_log.txt', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')


def find_index(lines, keyword, start=0):
    for i in range(start, len(lines)):
        if keyword.lower() in lines[i].lower():
            return i
    return None


def scrape_company_info(url, retries=3):
    attempt = 0
    while attempt < retries:
        try:
            response = requests.get(url, timeout=30, headers={'User-Agent': 'Mozilla/5.0'})
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')
                text = soup.get_text(separator='\n', strip=True)
                lines = text.split('\n')

                company_info_index = find_index(lines, "Registered office address") - 1
                company_info_line = lines[company_info_index]
                company_code = re.search(r'\((\d+)\)', company_info_line).group(1)
                company_name = company_info_line.split('for ')[1].split(' (')[0]
                company_link = f"https://find-and-update.company-information.service.gov.uk/company/{company_code}"

                next_accounts_index = find_index(lines, "Next accounts made up to")
                first_accounts_index = find_index(lines, "First accounts made up to")
                if next_accounts_index is not None:
                    account_due_index = find_index(lines, "due by", next_accounts_index)
                elif first_accounts_index is not None:
                    account_due_index = find_index(lines, "due by", first_accounts_index)
                else:
                    account_due_index = None

                next_statement_index = find_index(lines, "Next statement date")
                first_statement_index = find_index(lines, "First statement date")
                if next_statement_index is not None:
                    statement_due_index = find_index(lines, "due by", next_statement_index)
                elif first_statement_index is not None:
                    statement_due_index = find_index(lines, "due by", first_statement_index)
                else:
                    statement_due_index = None

                data_dict = {
                    "Company Code": company_code,
                    "Company": company_name,
                    "Registered office address": lines[
                        find_index(lines, "Registered office address") + 1] if find_index(lines,
                                                                                          "Registered office address") is not None else "N/A",
                    "Company status": lines[find_index(lines, "Company status") + 1] if find_index(lines,
                                                                                                   "Company status") is not None else "N/A",
                    "Company type": lines[find_index(lines, "Company type") + 1] if find_index(lines,
                                                                                               "Company type") is not None else "N/A",
                    "Incorporated on": lines[find_index(lines, "Incorporated on") + 1] if find_index(lines,
                                                                                                     "Incorporated on") is not None else "N/A",
                    "First accounts made up to": lines[
                        first_accounts_index + 1] if first_accounts_index is not None else "N/A",
                    "Next accounts made up to": lines[
                        next_accounts_index + 1] if next_accounts_index is not None else "N/A",
                    "Account Due By": lines[account_due_index + 1] if account_due_index is not None else "N/A",
                    "Last accounts made up to": lines[find_index(lines, "Last accounts made up to") + 1] if find_index(
                        lines, "Last accounts made up to") is not None else "N/A",
                    "First statement date": lines[
                        first_statement_index + 1] if first_statement_index is not None else "N/A",
                    "Next statement date": lines[
                        next_statement_index + 1] if next_statement_index is not None else "N/A",
                    "Statement Due By": lines[statement_due_index + 1] if statement_due_index is not None else "N/A",
                    "Last statement dated": lines[find_index(lines, "Last statement dated") + 1] if find_index(lines,
                                                                                                               "Last statement dated") is not None else "N/A",
                    "URL": company_link
                }

                sic_start_index = find_index(lines, "Nature of business (SIC)")
                if sic_start_index is not None:
                    end_keywords = ["Previous company names", "Tell us what you think of this service"]
                    end_indices = [find_index(lines, keyword, sic_start_index) for keyword in end_keywords]
                    end_indices = [index for index in end_indices if index is not None]
                    sic_end_index = min(end_indices) if end_indices else None

                    if sic_end_index is not None:
                        sic_entries = lines[sic_start_index + 1:sic_end_index]
                        data_dict["Nature of business (SIC)"] = ', '.join(sic_entries)
                    else:
                        data_dict["Nature of business (SIC)"] = "N/A"
                else:
                    data_dict["Nature of business (SIC)"] = "N/A"

                return data_dict
            else:
                logging.error(f"HTTP status code: {response.status_code} for URL: {url}")
        except requests.exceptions.RequestException as e:
            logging.error(f"Request failed: {e} for URL: {url}")
            time.sleep(2 ** attempt)  # Exponential backoff
            attempt += 1
    return None


def process_file():
    process_button.config(text="Processing your File...")
    root.update()

    file_path = filedialog.askopenfilename(title="Select Raw Excel File", filetypes=(("Excel files", "*.xlsx;*.xlsm"),))

    if not file_path:
        messagebox.showinfo("Info", "File selection cancelled.")
        process_button.config(text="Load Raw File")
        return

    try:
        df_urls = pd.read_excel(file_path, usecols=['Company Number'])
    except ValueError as ve:
        messagebox.showerror("Error",
                             "Invalid File Selected. Please select a valid Excel file with the required columns.")
        process_button.config(text="Load Raw File")
        return

    df_urls['Company Number'] = df_urls['Company Number'].apply(lambda x: str(x).zfill(8))
    df_urls['Generated URL'] = "https://find-and-update.company-information.service.gov.uk/company/" + df_urls[
        'Company Number']
    url_list = df_urls['Generated URL'].tolist()

    df_final = pd.DataFrame()
    invalid_companies = []
    erroneous_links = []

    for url in url_list:
        try:
            data_dict = scrape_company_info(url)
            if data_dict:
                df_final = pd.concat([df_final, pd.DataFrame([data_dict])], ignore_index=True)
            else:
                invalid_companies.append(url)
        except Exception as e:
            erroneous_links.append(url)
            logging.error(f"Error processing {url}: {e}")

    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx;*.xlsm")],
                                             initialfile=f"scraped_company_info_{datetime.now().strftime('%Y-%m-%d-%H-%M')}.xlsx")

    if not save_path:
        messagebox.showinfo("Info", "Save operation cancelled.")
        process_button.config(text="Load Raw File")
        return

    date_columns = [
        "Incorporated on", "First accounts made up to", "Next accounts made up to", "Account Due By",
        "Last accounts made up to", "First statement date", "Next statement date", "Statement Due By",
        "Last statement dated"
    ]

    for col in date_columns:
        df_final[col] = pd.to_datetime(df_final[col], errors='coerce', format='%d %B %Y')

    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name='Scraped', index=False)

        if invalid_companies:
            df_invalid = pd.DataFrame(invalid_companies, columns=['Invalid Links'])
            df_invalid.to_excel(writer, sheet_name='Invalid Links', index=False)

        if erroneous_links:
            df_erroneous = pd.DataFrame(erroneous_links, columns=['Erroneous Links'])
            df_erroneous.to_excel(writer, sheet_name='Erroneous Links', index=False)

        workbook = writer.book
        sheet = writer.sheets['Scraped']

        # Apply cell styles to date columns
        date_style = NamedStyle(name='date_style', number_format='DD-MM-YYYY')
        date_style.font = Font(size=10)
        date_style.fill = PatternFill(fill_type='solid', fgColor='FFFFCC')
        workbook.add_named_style(date_style)

        for column in date_columns:
            col_idx = df_final.columns.get_loc(column) + 1
            for row in range(2, len(df_final) + 2):
                cell = sheet.cell(row=row, column=col_idx)
                cell.style = date_style

        # Add header color and font color
        header_fill = PatternFill(start_color='4D9BE9', end_color='4D9BE9', fill_type='solid')
        header_font = Font(color='FFFFFF')

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

            # Adjust column width to fit the text
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    messagebox.showinfo("Success", f"Data successfully saved to {save_path}")
    process_button.config(text="Load Raw File")


def download_log():
    try:
        save_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
        if save_path:
            with open('scrapify_log.txt', 'r') as src:
                with open(save_path, 'w') as dst:
                    dst.write(src.read())
            messagebox.showinfo("Success", "Log file downloaded successfully.")
        else:
            messagebox.showinfo("Cancelled", "Download cancelled.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to download log file: {e}")


# Set up GUI
root = Tk()
root.title("Scrapify V1.0")
root.geometry("300x150")

process_button = Button(root, text="Load Raw File", command=process_file, width=20, height=2)
process_button.pack(pady=10)

download_button = Button(root, text="Download Log File", command=download_log, width=20, height=2)
download_button.pack(pady=10)

root.mainloop()

